import telebot
import openpyxl

bot = telebot.TeleBot('7351727427:AAE-xAhU7CIc453vIuVM0kUMuX0FUJO4hKc')


@bot.message_handler(commands=['start'])
def start(message):
    bot.send_message(message.chat.id, 'Привет! Отправьте мне файл Excel для дальнейшей работы.')


@bot.message_handler(content_types=['document'])
def handle_document(message):
    if message.document.mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)

        with open('Отчет по студентам.xlsx', 'wb') as new_file:
            new_file.write(downloaded_file)

        bot.send_message(message.chat.id, 'Файл успешно загружен. Вы можете использовать команды /command1 и /command2.')
    else:
        bot.send_message(message.chat.id, 'Пожалуйста, загрузите файл Excel формата .xlsx')


@bot.message_handler(commands=['command1'])
def check_homework(message):
    wb = openpyxl.load_workbook('Отчет по студентам.xlsx')
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=20):
        student_name = row[0].value
        homework_percentage = row[19].value

        if homework_percentage < 50:
            bot.send_message(message.chat.id,
                             f'{student_name} выполнил только {homework_percentage}% ДЗ за последние 30 дней.')

    bot.send_message(message.chat.id, 'Проверка пройдена.')


@bot.message_handler(commands=['command2'])
def check_student(message):
    wb = openpyxl.load_workbook('Отчет по студентам.xlsx')
    sheet = wb.active

    student_messages = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=20):
        student_name = row[0].value
        average_score = row[17].value

        if isinstance(average_score, (int, float)) and average_score < 3:
            student_messages.append(f'{student_name}, средний балл: {average_score}.')

    if student_messages:
        bot.send_message(message.chat.id, '\n'.join(student_messages))
    else:
        bot.send_message(message.chat.id, 'Все студенты имеют средний балл выше 3.')

    bot.send_message(message.chat.id, 'Поиск студентов прошёл успешно.')


bot.polling(none_stop=True, interval=0)


